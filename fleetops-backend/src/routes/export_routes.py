"""
export_routes.py — Endpoints de exportación con formato profesional.

Genera archivos Excel (.xlsx) y CSV con:
  • Encabezado corporativo (Tersa Mundi)
  • Tabla de datos formateada con colores alternados
  • Gráfica de barras nativa de Excel
  • Hoja de resumen ejecutivo

Endpoints:
  GET /api/export/excel?report=<tipo>&year=2026&month=5
  GET /api/export/csv?report=<tipo>&year=2026&month=5

Tipos de reporte disponibles:
  costo-combustible  | km-rendimiento  | comparativo
  operatividad       | tonelaje        | diesel-diario
  bascula-mensual    | bascula-turnos
"""

import csv
import io
from datetime import date as _date

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from src.services import state_store
from src.database import db_service
from src.routes.report_routes import _current, _get_all_unit_metrics

router = APIRouter(prefix="/api/export", tags=["Exportación"])

# ── Constantes de estilo ──────────────────────────────────────────────────────
BLUE       = "2563EB"
BLUE_DARK  = "1E3A5F"
BLUE_LIGHT = "EFF6FF"
BLUE_MID   = "DBEAFE"
WHITE      = "FFFFFF"
GRAY       = "6B7280"
GRAY_LIGHT = "F9FAFB"
GREEN      = "16A34A"
AMBER      = "D97706"
RED_COL    = "DC2626"

EMPRESA = "Tersa Mundi S.A. de C.V."
SISTEMA = "FleetOps"

THIN   = Side(style="thin",   color="D1D5DB")
MEDIUM = Side(style="medium", color=BLUE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_HEADER = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

LABELS = {
    "eco":                 "No. Económico",
    "km_total":            "Km Total",
    "km_prod":             "Km Productivo",
    "km_trasl":            "Km Traslado",
    "km_disp":             "Km Disposición",
    "km_no_prod":          "Km No Productivo",
    "litros":              "Litros",
    "km_por_litro":        "Km/Lt",
    "alerta_rendimiento":  "Alerta",
    "diesel_lts":          "Lts Diesel",
    "diesel_importe":      "Importe Diesel",
    "gasolina_lts":        "Lts Gasolina",
    "total_lts":           "Total Litros",
    "total_importe":       "Total Importe",
    "precio_diesel":       "Precio/Lt",
    "toneladas":           "Toneladas",
    "num_viajes":          "Viajes",
    "tons_prom_por_viaje": "Ton/Viaje Prom.",
    "tons_ticket":         "Ton Ticket",
    "variacion_pct":       "Variación %",
    "t_prod_hrs":          "T. Productivo (hrs)",
    "t_no_prod_hrs":       "T. No Productivo (hrs)",
    "ton_por_t_prod":      "Ton/Hr Productiva",
    "ton_por_km_prod":     "Ton/Km Productivo",
    "pct_carga":           "% Carga",
    "periodo":             "Período",
    "litros_mes":          "Lts Mes",
    "km_mes":              "Km Mes",
    "tipo_cliente":        "Tipo Cliente",
    "tipo_residuo":        "Tipo Residuo",
    "viajes":              "Viajes",
    "promedio_ton_viaje":  "Ton/Viaje Prom.",
    "pct_toneladas":       "% del Total",
    "turno":               "Turno",
    "dia":                 "Día",
    "importe":             "Importe",
    "num_eco":             "No. Económico",
}

HIDDEN = {"vehicle_id"}

NUM_FORMAT = {
    "km_total":            "#,##0",
    "km_prod":             "#,##0",
    "km_trasl":            "#,##0",
    "km_disp":             "#,##0",
    "km_no_prod":          "#,##0",
    "km_mes":              "#,##0",
    "km_por_litro":        "0.000",
    "litros":              "#,##0.0",
    "litros_mes":          "#,##0.0",
    "diesel_lts":          "#,##0.0",
    "gasolina_lts":        "#,##0.0",
    "total_lts":           "#,##0.0",
    "diesel_importe":      '"$"#,##0.00',
    "total_importe":       '"$"#,##0.00',
    "precio_diesel":       '"$"#,##0.00',
    "importe":             '"$"#,##0.00',
    "pct_carga":           "0.0%",
    "pct_toneladas":       "0.0%",
    "variacion_pct":       "0.0%",
    "toneladas":           "#,##0.000",
    "tons_prom_por_viaje": "#,##0.000",
    "tons_ticket":         "#,##0.000",
    "promedio_ton_viaje":  "#,##0.000",
}


# ── Helpers de estilo ─────────────────────────────────────────────────────────

def _style(cell, bold=False, color="000000", bg=None,
           align="left", size=9, border=False, wrap=False, italic=False):
    cell.font      = Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    if border:
        cell.border = BORDER


def _header_row(ws, row: int, n_cols: int, text: str, size: int = 13):
    ws.row_dimensions[row].height = 28
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    _style(c, bold=True, color=WHITE, bg=BLUE, size=size)


def _period_label(year: int, month: int) -> str:
    meses = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
    }
    return f"{meses.get(month, str(month))} de {year}"


def _alert_color(val) -> str | None:
    if not val:
        return None
    v = str(val).lower()
    if v == "critico":
        return RED_COL
    if v == "bajo":
        return AMBER
    if v == "normal":
        return GREEN
    return None


# ── Constructor de Excel genérico ─────────────────────────────────────────────

def _build_excel(
    rows:       list[dict],
    title:      str,
    period:     str,
    unit_count: int,
    totals:     dict = None,
    chart_col:  str  = None,
    extra_sheets: list[tuple[str, list[dict]]] = None,  # [(nombre, filas), ...]
) -> bytes:
    wb = Workbook()

    # ══ HOJA DE DATOS ════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Datos"
    ws.sheet_view.showGridLines = False

    keys  = [k for k in (rows[0].keys() if rows else []) if k not in HIDDEN]
    heads = [LABELS.get(k, k.replace("_", " ").title()) for k in keys]
    n     = len(keys)

    # Encabezado corporativo
    _header_row(ws, 1, n, EMPRESA)
    ws.row_dimensions[2].height = 20
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    c2 = ws.cell(row=2, column=1, value=title)
    _style(c2, bold=True, color=BLUE_DARK, size=11)

    ws.row_dimensions[3].height = 17
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n)
    c3 = ws.cell(row=3, column=1, value=period)
    _style(c3, color=GRAY, size=10)

    ws.row_dimensions[4].height = 15
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n)
    c4 = ws.cell(row=4, column=1, value=f"{unit_count} unidades · {SISTEMA}")
    _style(c4, color=GRAY, size=8, italic=True)

    ws.row_dimensions[5].height = 8

    # Cabeceras de tabla
    ws.row_dimensions[6].height = 22
    for ci, label in enumerate(heads, 1):
        c = ws.cell(row=6, column=ci, value=label)
        _style(c, bold=True, color=WHITE, bg=BLUE, align="center", size=9, border=True)

    # Filas de datos
    for ri, row in enumerate(rows):
        excel_row = 7 + ri
        ws.row_dimensions[excel_row].height = 17
        bg = BLUE_LIGHT if ri % 2 == 0 else WHITE
        for ci, key in enumerate(keys, 1):
            val = row.get(key)
            c   = ws.cell(row=excel_row, column=ci, value=val)
            is_text = ci == 1
            _style(c, bg=bg, align="left" if is_text else "center", size=9, border=True)
            if key in NUM_FORMAT and val is not None:
                c.number_format = NUM_FORMAT[key]
            # Colorear alertas
            if key == "alerta_rendimiento":
                ac = _alert_color(val)
                if ac:
                    c.font = Font(name="Arial", bold=True, color=ac, size=9)

    last_data = 6 + len(rows)

    # Fila de TOTAL
    total_row = last_data + 1
    ws.row_dimensions[total_row].height = 20
    c_tot = ws.cell(row=total_row, column=1, value="TOTAL")
    _style(c_tot, bold=True, bg=BLUE_MID, align="left", size=9, border=True)
    for ci, key in enumerate(keys[1:], 2):
        c = ws.cell(row=total_row, column=ci)
        col_letter = get_column_letter(ci)
        sample = next((r.get(key) for r in rows if r.get(key) is not None), None)
        if isinstance(sample, (int, float)):
            c.value = f"=SUM({col_letter}7:{col_letter}{last_data})"
            if key in NUM_FORMAT:
                c.number_format = NUM_FORMAT[key]
        else:
            c.value = ""
        _style(c, bold=True, bg=BLUE_MID, align="center", size=9, border=True)

    # Anchos de columna
    for ci, key in enumerate(keys, 1):
        max_len = max(
            len(heads[ci - 1]),
            max((len(str(r.get(key, ""))) for r in rows), default=0),
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(32, max(10, max_len + 3))

    # Gráfica de barras
    chart_key = chart_col or _detect_chart_key(keys)
    if chart_key and chart_key in keys:
        chart_ci = keys.index(chart_key) + 1
        eco_ci   = keys.index("eco") + 1 if "eco" in keys else (keys.index("num_eco") + 1 if "num_eco" in keys else 1)

        chart           = BarChart()
        chart.type      = "col"
        chart.grouping  = "clustered"
        chart.title     = f"{LABELS.get(chart_key, chart_key)} por Unidad"
        chart.y_axis.title = LABELS.get(chart_key, chart_key)
        chart.x_axis.title = "Unidad"
        chart.style     = 10
        chart.width     = 28
        chart.height    = 14

        data_ref = Reference(ws, min_col=chart_ci, min_row=6, max_row=last_data)
        cats_ref = Reference(ws, min_col=eco_ci,   min_row=7, max_row=last_data)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"A{total_row + 3}")

    # ══ HOJAS EXTRA (para diesel-diario, bascula-turnos, etc.) ════════════════
    if extra_sheets:
        for sheet_name, sheet_rows in extra_sheets:
            if not sheet_rows:
                continue
            wse = wb.create_sheet(title=sheet_name[:31])
            wse.sheet_view.showGridLines = False
            s_keys  = [k for k in sheet_rows[0].keys() if k not in HIDDEN]
            s_heads = [LABELS.get(k, k.replace("_", " ").title()) for k in s_keys]
            sn      = len(s_keys)

            _header_row(wse, 1, sn, f"{EMPRESA} — {sheet_name}")
            wse.row_dimensions[2].height = 17
            wse.merge_cells(start_row=2, start_column=1, end_row=2, end_column=sn)
            wse.cell(row=2, column=1, value=period)
            _style(wse.cell(row=2, column=1), color=GRAY, size=9)
            wse.row_dimensions[3].height = 8

            for ci, label in enumerate(s_heads, 1):
                c = wse.cell(row=4, column=ci, value=label)
                _style(c, bold=True, color=WHITE, bg=BLUE, align="center", size=9, border=True)
            wse.row_dimensions[4].height = 20

            for ri, row in enumerate(sheet_rows):
                er = 5 + ri
                bg = BLUE_LIGHT if ri % 2 == 0 else WHITE
                wse.row_dimensions[er].height = 17
                for ci, key in enumerate(s_keys, 1):
                    val = row.get(key)
                    c   = wse.cell(row=er, column=ci, value=val)
                    _style(c, bg=bg, align="left" if ci == 1 else "center", size=9, border=True)
                    if key in NUM_FORMAT and val is not None:
                        c.number_format = NUM_FORMAT[key]

            for ci, key in enumerate(s_keys, 1):
                max_len = max(len(s_heads[ci-1]),
                              max((len(str(r.get(key,""))) for r in sheet_rows), default=0))
                wse.column_dimensions[get_column_letter(ci)].width = min(32, max(10, max_len + 3))

    # ══ HOJA RESUMEN ══════════════════════════════════════════════════════════
    wsr = wb.create_sheet(title="Resumen")
    wsr.sheet_view.showGridLines = False
    wsr.column_dimensions["A"].width = 30
    wsr.column_dimensions["B"].width = 22

    _header_row(wsr, 1, 2, EMPRESA)
    wsr.row_dimensions[2].height = 20
    wsr.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    _style(wsr.cell(row=2, column=1, value=f"Resumen — {title}"), bold=True, color=BLUE_DARK, size=11)
    wsr.row_dimensions[3].height = 17
    wsr.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    _style(wsr.cell(row=3, column=1, value=period), color=GRAY, size=10)
    wsr.row_dimensions[4].height = 8

    summary_items = [
        ("Empresa",          EMPRESA),
        ("Período",          period),
        ("Unidades",         unit_count),
        ("Sistema",          SISTEMA),
        ("Generado",         str(_date.today())),
    ]
    if totals:
        summary_items.append(("", ""))
        summary_items.append(("TOTALES DEL REPORTE", ""))
        for k, v in totals.items():
            summary_items.append((LABELS.get(k, k), v))

    for sri, (label, value) in enumerate(summary_items, 5):
        wsr.row_dimensions[sri].height = 18
        cl = wsr.cell(row=sri, column=1, value=label)
        cv = wsr.cell(row=sri, column=2, value=value)
        bg = BLUE_LIGHT if sri % 2 == 0 else WHITE
        if label in ("TOTALES DEL REPORTE", ""):
            bg = BLUE_MID
            _style(cl, bold=True, bg=bg, size=9)
            _style(cv, bold=True, bg=bg, size=9, align="right")
        else:
            _style(cl, bg=bg, size=9)
            _style(cv, bg=bg, size=9, align="right")

    last_s = 5 + len(summary_items) + 2
    _style(wsr.cell(row=last_s, column=1, value=f"Generado por {SISTEMA}"),
           color=GRAY, size=8, italic=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _detect_chart_key(keys: list[str]) -> str | None:
    for k in ["total_lts", "litros", "diesel_lts", "km_total", "toneladas"]:
        if k in keys:
            return k
    return None


# ── CSV builder ───────────────────────────────────────────────────────────────

def _build_csv(rows: list[dict]) -> bytes:
    if not rows:
        return b""
    buf = io.StringIO()
    keys  = [k for k in rows[0].keys() if k not in HIDDEN]
    heads = [LABELS.get(k, k.replace("_", " ").title()) for k in keys]
    writer = csv.writer(buf)
    writer.writerow(heads)
    for row in rows:
        writer.writerow([row.get(k, "") for k in keys])
    return buf.getvalue().encode("utf-8-sig")  # BOM para Excel en Windows


# ── Helpers de datos ──────────────────────────────────────────────────────────

async def _data_costo_combustible(year: int, month: int):
    from datetime import datetime, date
    today = date.today()

    def _in_month(record, y, m):
        fecha = record.get("fecha")
        if not fecha:
            return False
        try:
            dt = datetime.fromisoformat(str(fecha).replace("Z", "")) if isinstance(fecha, str) else fecha
            return dt.year == y and dt.month == m
        except Exception:
            return False

    vehicles = state_store.get_vehicles()
    rows, total_lts_g, total_imp_g = [], 0.0, 0.0

    # Meses pasados: los registros ya no viven en memoria (state_store solo
    # carga el mes actual al arrancar), así que se consultan directo en MySQL.
    is_current_month = (year == today.year and month == today.month)
    fuel_by_vehicle: dict[str, list[dict]] = {}
    if not is_current_month:
        try:
            month_records = await db_service.get_fuel_records_by_month(year, month)
        except Exception:
            month_records = []
        for r in month_records:
            fuel_by_vehicle.setdefault(str(r.get("vehicle_id", "")), []).append(r)

    for v in vehicles:
        vid = str(v.get("ras_vei_id", ""))
        eco = v.get("ras_vei_eco") or v.get("ras_vei_placa") or vid
        all_fuel   = state_store.get_fuel_records_for_vehicle(vid) if is_current_month else fuel_by_vehicle.get(vid, [])
        month_recs = [r for r in all_fuel if _in_month(r, year, month)]
        met = state_store.get_unit_metrics(vid) if (year == today.year and month == today.month) else {}
        if not met:
            hist = state_store.get_metrics_historico(year, month)
            met  = next((h for h in hist if str(h.get("vehicle_id")) == vid), {})

        diesel_lts = met.get("diesel_lts") or sum(r.get("liters", 0) for r in month_recs if r.get("tipo") in ("DIESEL", None, ""))
        gas_lts    = met.get("gasolina_lts", 0) or sum(r.get("liters", 0) for r in month_recs if r.get("tipo") in ("GASOLINA_COMUN", "GASOLINA_PREMIUM"))
        total_lts  = round(diesel_lts + gas_lts, 2)
        if total_lts <= 0:
            continue

        total_imp  = met.get("total_importe") or round(sum(r.get("liters", 0) * r.get("price_per_liter", 0) for r in month_recs), 2)
        precio     = met.get("precio_diesel") or (month_recs[-1].get("price_per_liter") if month_recs else 0)
        rows.append({
            "eco":            eco,
            "diesel_lts":     round(diesel_lts, 2),
            "diesel_importe": met.get("diesel_importe") or round(diesel_lts * (precio or 0), 2),
            "gasolina_lts":   round(gas_lts, 2),
            "total_lts":      total_lts,
            "total_importe":  total_imp,
            "precio_diesel":  precio,
        })
        total_lts_g += total_lts
        total_imp_g += total_imp

    totals = {"total_lts": round(total_lts_g, 2), "total_importe": round(total_imp_g, 2)}
    return rows, totals


def _data_km_rendimiento(year: int, month: int) -> list[dict]:
    metrics = _get_all_unit_metrics(year, month)
    return [
        {
            "eco":                m.get("eco"),
            "km_prod":            m.get("km_prod"),
            "km_trasl":           m.get("km_trasl"),
            "km_disp":            m.get("km_disp"),
            "km_total":           m.get("km_total"),
            "litros":             m.get("diesel_lts"),
            "km_por_litro":       m.get("km_por_litro"),
            "alerta_rendimiento": m.get("alerta_rendimiento"),
        }
        for m in metrics
    ]


def _data_operatividad(year: int, month: int) -> list[dict]:
    metrics = _get_all_unit_metrics(year, month)
    return [
        {
            "eco":             m.get("eco"),
            "toneladas":       m.get("toneladas"),
            "t_prod_hrs":      m.get("t_prod_hrs"),
            "t_no_prod_hrs":   m.get("t_no_prod_hrs"),
            "ton_por_t_prod":  m.get("ton_por_t_prod"),
            "km_prod":         m.get("km_prod"),
            "km_no_prod":      m.get("km_no_prod"),
            "ton_por_km_prod": m.get("ton_por_km_prod"),
            "pct_carga":       m.get("pct_carga"),
        }
        for m in metrics
    ]


def _data_tonelaje(year: int, month: int) -> list[dict]:
    metrics = _get_all_unit_metrics(year, month)
    return [
        {
            "eco":                 m.get("eco"),
            "toneladas":           m.get("toneladas"),
            "num_viajes":          m.get("num_viajes"),
            "tons_prom_por_viaje": m.get("tons_prom_por_viaje"),
            "tons_ticket":         m.get("tons_ticket"),
            "variacion_pct":       m.get("variacion_pct"),
        }
        for m in metrics
    ]


def _data_comparativo() -> list[dict]:
    today = _date.today()
    rows  = []
    for i in range(3):
        mo = today.month - i
        yr = today.year
        if mo <= 0:
            mo += 12
            yr -= 1
        metrics = _get_all_unit_metrics(yr, mo)
        for m in metrics:
            rows.append({
                "eco":           m.get("eco"),
                "periodo":       _period_label(yr, mo),
                "km_total":      m.get("km_total"),
                "litros":        m.get("diesel_lts"),
                "km_por_litro":  m.get("km_por_litro"),
                "total_importe": m.get("total_importe"),
            })
    return rows


async def _data_bascula_mensual(year: int, month: int):
    from src.database import db_service
    por_unidad = await db_service.get_bascula_monthly_by_eco(year, month)

    from collections import defaultdict
    eco_agg: dict[str, dict] = defaultdict(lambda: {"toneladas": 0.0, "viajes": 0})
    for row in por_unidad:
        eco = row["num_eco"]
        eco_agg[eco]["toneladas"] += row["toneladas"]
        eco_agg[eco]["viajes"]    += row["viajes"]

    rows = [
        {
            "num_eco":            eco,
            "toneladas":          round(d["toneladas"], 3),
            "viajes":             d["viajes"],
            "promedio_ton_viaje": round(d["toneladas"] / d["viajes"], 3) if d["viajes"] else 0,
        }
        for eco, d in sorted(eco_agg.items(), key=lambda x: -x[1]["toneladas"])
    ]
    total_tons   = round(sum(r["toneladas"] for r in rows), 3)
    total_viajes = sum(r["viajes"] for r in rows)
    totals = {"toneladas": total_tons, "viajes": total_viajes}
    return rows, totals


async def _data_bascula_turnos(year: int, month: int):
    from src.database import db_service
    rows_raw = await db_service.get_bascula_by_turno(year, month)
    # Aplanar: una fila por (dia, turno, tipo_cliente)
    return [
        {
            "dia":          r["dia"],
            "turno":        r["turno"],
            "tipo_cliente": r["tipo_cliente"],
            "toneladas":    r["toneladas"],
            "viajes":       r["viajes"],
        }
        for r in rows_raw
    ]


async def _data_diesel_diario(year: int, month: int):
    from src.database import db_service
    raw = await db_service.get_diesel_daily_by_vehicle(year, month)
    vehicles = state_store.get_vehicles()
    id_to_eco = {str(v["ras_vei_id"]): v.get("ras_vei_eco") or v.get("ras_vei_placa", "") for v in vehicles}

    rows = [
        {
            "eco":     id_to_eco.get(r["vehicle_id"], r["vehicle_id"]),
            "dia":     r["fecha"],
            "litros":  r["litros"],
            "importe": r["importe"],
        }
        for r in raw
    ]
    total_lts = round(sum(r["litros"] for r in rows), 2)
    total_imp = round(sum(r["importe"] for r in rows), 2)
    totals    = {"litros": total_lts, "importe": total_imp}
    return rows, totals


# ── Endpoints ─────────────────────────────────────────────────────────────────

_REPORT_TITLES = {
    "costo-combustible": "Consumo y Costo de Combustible por Unidad",
    "km-rendimiento":    "Kilómetros Recorridos y Rendimiento de Combustible",
    "operatividad":      "Operatividad por Unidad",
    "tonelaje":          "Tonelaje por Unidad",
    "comparativo":       "Reporte Comparativo — Últimos 3 Meses",
    "bascula-mensual":   "Báscula — Resumen Mensual por Unidad",
    "bascula-turnos":    "Báscula — Tonelaje por Turno y Tipo Cliente",
    "diesel-diario":     "Consumo de Diesel por Día y Unidad",
}

_VALID_REPORTS = set(_REPORT_TITLES.keys())


async def _resolve_data(report: str, year: int, month: int):
    """
    Retorna (rows, title, totals, extra_sheets).
    extra_sheets = [(nombre_hoja, filas)] para reportes multi-hoja.
    """
    if report == "costo-combustible":
        rows, totals = await _data_costo_combustible(year, month)
        return rows, _REPORT_TITLES[report], totals, None

    if report == "km-rendimiento":
        rows = _data_km_rendimiento(year, month)
        return rows, _REPORT_TITLES[report], {}, None

    if report == "operatividad":
        rows = _data_operatividad(year, month)
        return rows, _REPORT_TITLES[report], {}, None

    if report == "tonelaje":
        rows = _data_tonelaje(year, month)
        total_tons = round(sum(r.get("toneladas") or 0 for r in rows), 3)
        return rows, _REPORT_TITLES[report], {"toneladas": total_tons}, None

    if report == "comparativo":
        rows = _data_comparativo()
        return rows, _REPORT_TITLES[report], {}, None

    if report == "bascula-mensual":
        rows, totals = await _data_bascula_mensual(year, month)
        return rows, _REPORT_TITLES[report], totals, None

    if report == "bascula-turnos":
        rows = await _data_bascula_turnos(year, month)
        return rows, _REPORT_TITLES[report], {}, None

    if report == "diesel-diario":
        rows, totals = await _data_diesel_diario(year, month)
        # Extra: pivot por vehículo como segunda hoja
        from collections import defaultdict
        pivot: dict[str, dict] = defaultdict(lambda: {"litros": 0.0, "importe": 0.0})
        for r in rows:
            pivot[r["eco"]]["litros"]  += r["litros"]
            pivot[r["eco"]]["importe"] += r["importe"]
        km_data = [
            {"eco": eco, "litros_mes": round(d["litros"], 2), "importe": round(d["importe"], 2)}
            for eco, d in sorted(pivot.items())
        ]
        return rows, _REPORT_TITLES[report], totals, [("Resumen por Unidad", km_data)]

    # Fallback genérico
    metrics = _get_all_unit_metrics(year, month)
    rows = [dict(m) for m in metrics]
    return rows, report.replace("-", " ").title(), {}, None


@router.get("/excel")
async def export_excel(
    report: str = Query(..., description=f"Tipo de reporte: {', '.join(_VALID_REPORTS)}"),
    year:   int = Query(default=None),
    month:  int = Query(default=None),
    units:  str = Query(default=None, description="Ecos separados por coma para filtrar"),
):
    """
    Genera y descarga un Excel profesional con encabezado corporativo y gráfica de barras.

    Ejemplos:
      GET /api/export/excel?report=costo-combustible&year=2026&month=5
      GET /api/export/excel?report=bascula-mensual&month=5
      GET /api/export/excel?report=diesel-diario&units=TM-1,TM-2
    """
    y, m = _current(year, month)
    period = _period_label(y, m)
    selected_units = [u.strip() for u in units.split(",")] if units else None

    rows, title, totals, extra_sheets = await _resolve_data(report, y, m)

    if selected_units:
        eco_key = "eco" if rows and "eco" in rows[0] else "num_eco"
        rows = [r for r in rows if r.get(eco_key) in selected_units]

    if not rows:
        raise HTTPException(status_code=404, detail="Sin datos para los parámetros indicados.")

    xlsx_bytes = _build_excel(
        rows=rows,
        title=title,
        period=period,
        unit_count=len(rows),
        totals=totals or {},
        extra_sheets=extra_sheets,
    )

    filename = f"{report}_{y}-{m:02d}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/csv")
async def export_csv(
    report: str = Query(..., description=f"Tipo de reporte: {', '.join(_VALID_REPORTS)}"),
    year:   int = Query(default=None),
    month:  int = Query(default=None),
    units:  str = Query(default=None, description="Ecos separados por coma para filtrar"),
):
    """
    Genera y descarga un CSV plano, compatible con Excel (BOM UTF-8).

    Ejemplos:
      GET /api/export/csv?report=bascula-mensual&month=5
      GET /api/export/csv?report=km-rendimiento
    """
    y, m = _current(year, month)
    selected_units = [u.strip() for u in units.split(",")] if units else None

    rows, title, _totals, _extra = await _resolve_data(report, y, m)

    if selected_units:
        eco_key = "eco" if rows and "eco" in rows[0] else "num_eco"
        rows = [r for r in rows if r.get(eco_key) in selected_units]

    if not rows:
        raise HTTPException(status_code=404, detail="Sin datos para los parámetros indicados.")

    csv_bytes = _build_csv(rows)
    filename  = f"{report}_{y}-{m:02d}.csv"
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/reportes-disponibles")
async def reportes_disponibles():
    """Lista todos los tipos de reporte disponibles para exportar."""
    return {
        "formatos": ["excel", "csv"],
        "reportes": [
            {"id": rid, "titulo": title, "endpoint_excel": f"/api/export/excel?report={rid}", "endpoint_csv": f"/api/export/csv?report={rid}"}
            for rid, title in _REPORT_TITLES.items()
        ],
    }